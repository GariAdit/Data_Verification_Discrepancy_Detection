import pandas as pd
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import warnings
from typing import List, Dict, Tuple, Optional

class InvoiceComparator:
    def __init__(self):
        self.invoice_data = []
        self.master_data = None
        self.discrepancies = []
        
    def extract_invoice_data(self, pdf_path: str) -> List[Dict]:
        """
        Extract invoice data from PDF file
        """
        data = []
        
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                # Extract tables from the page
                tables = page.extract_tables()
                
                for table in tables:
                    # Find the header row (the one with 'Item' in it)
                    header_row_idx = None
                    for i, row in enumerate(table):
                        if row and 'Item' in str(row[0]):
                            header_row_idx = i
                            break
                    
                    if header_row_idx is not None:
                        headers = [h.strip().lower() if h else '' for h in table[header_row_idx]]
                        
                        # Process data rows
                        for row in table[header_row_idx + 1:]:
                            if not row or not any(row):
                                continue
                            
                            # Create a dictionary for each item
                            item_data = {}
                            for i, header in enumerate(headers):
                                if i < len(row) and header:
                                    try:
                                        # Clean the value
                                        value = str(row[i]).strip()
                                        if value.replace('.', '', 1).isdigit():
                                            value = float(value)
                                        item_data[header] = value
                                    except (ValueError, IndexError):
                                        continue
                            
                            if 'item' in item_data and item_data['item'] and 'total' in item_data:
                                data.append(item_data)
        
        return data
    
    def load_master_data(self, excel_path: str) -> pd.DataFrame:
        """
        Load master data from Excel file
        """
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            df = pd.read_excel(excel_path, engine='openpyxl')
        
        # Clean column names
        df.columns = [str(col).strip().lower() for col in df.columns]
        
        # Convert item names to string and strip whitespace
        if 'item' in df.columns:
            df['item'] = df['item'].astype(str).str.strip()
        
        return df
    
    def compare_data(self, invoice_data: List[Dict], master_df: pd.DataFrame) -> List[Dict]:
        """
        Compare invoice data with master data and identify discrepancies
        """
        discrepancies = []
        
        for item in invoice_data:
            invoice_item = str(item.get('item', '')).strip()
            invoice_total = item.get('total', 0)
            
            if isinstance(invoice_total, str):
                # Clean total value (remove currency symbols, commas, etc.)
                invoice_total = ''.join(c for c in invoice_total if c.isdigit() or c == '.')
                try:
                    invoice_total = float(invoice_total)
                except ValueError:
                    invoice_total = 0
            
            # Find matching item in master data
            master_item = master_df[master_df['item'].str.strip().str.lower() == invoice_item.lower()]
            
            if not master_item.empty:
                master_total = master_item.iloc[0].get('total', 0)
                
                # Compare totals
                if abs(invoice_total - master_total) > 0.01:  # Allow for small rounding differences
                    discrepancies.append({
                        'item': invoice_item,
                        'total_price_in_invoice': invoice_total,
                        'total_price_in_master_data': master_total,
                        'discrepancy': invoice_total - master_total,
                        'discrepancy_percentage': abs((invoice_total - master_total) / master_total * 100) 
                            if master_total != 0 else float('inf')
                    })
            else:
                discrepancies.append({
                    'item': invoice_item,
                    'total_price_in_invoice': invoice_total,
                    'total_price_in_master_data': 'Not found in master data',
                    'discrepancy': 'N/A',
                    'discrepancy_percentage': 'N/A'
                })
        
        return discrepancies
    
    def generate_report(self, discrepancies: List[Dict], output_path: str) -> None:
        """
        Generate an Excel report with discrepancies
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Discrepancies Report"
        
        # Headers
        headers = [
            "Item", 
            "Total Price in Invoice", 
            "Total Price in Master Data", 
            "Discrepancy Amount",
            "Discrepancy Percentage (%)"
        ]
        
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Add data
        for item in discrepancies:
            ws.append([
                item['item'],
                item['total_price_in_invoice'],
                item['total_price_in_master_data'],
                item['discrepancy'],
                item['discrepancy_percentage']
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Format numbers
        for row in ws.iter_rows(min_row=2, max_col=5, max_row=len(discrepancies)+1):
            for i, cell in enumerate(row):
                if i in (1, 2, 3) and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                elif i == 4 and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00"%"'
        
        # Highlight significant discrepancies
        for row in ws.iter_rows(min_row=2, max_col=5, max_row=len(discrepancies)+1):
            discrepancy_cell = row[3]
            if isinstance(discrepancy_cell.value, (int, float)) and abs(discrepancy_cell.value) > 0.01:
                if discrepancy_cell.value > 0:
                    discrepancy_cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                else:
                    discrepancy_cell.fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
        
        wb.save(output_path)
    
    def process_invoices(self, invoice_paths: List[str], master_data_path: str, output_path: str) -> None:
        """
        Process all invoices and generate a consolidated report
        """
        all_discrepancies = []
        
        # Load master data once
        self.master_data = self.load_master_data(master_data_path)
        
        for invoice_path in invoice_paths:
            # Extract data from invoice
            invoice_data = self.extract_invoice_data(invoice_path)
            
            # Compare with master data
            discrepancies = self.compare_data(invoice_data, self.master_data)
            all_discrepancies.extend(discrepancies)
        
        # Generate report
        if all_discrepancies:
            self.generate_report(all_discrepancies, output_path)
            print(f"Report generated successfully at {output_path}")
        else:
            print("No discrepancies found between invoices and master data.")

